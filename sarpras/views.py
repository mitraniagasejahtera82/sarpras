
# =========================================================
# DJANGO CORE
# =========================================================
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib import messages
from django.utils import timezone
from django.db import transaction
from django.db.models import Q, Sum, Count
from django.core.paginator import Paginator
from django.template.loader import get_template, render_to_string
from django.contrib.staticfiles import finders
from django.conf import settings

# =========================================================
# PYTHON STANDARD LIBRARY
# =========================================================
import os
import json
import csv
import calendar
import base64
from io import BytesIO, TextIOWrapper
from datetime import date

# =========================================================
# THIRD PARTY LIBRARY
# =========================================================
from xhtml2pdf import pisa
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import qrcode
from django.db.models.functions import TruncMonth


# =========================================================
# LOCAL MODELS
# =========================================================
from .models import (
    Tanah,
    PeralatanMesin,
    Gedung,
    Ruangan,
    Jalan,
    Buku,
    Peminjaman,
    BarangHabisPakai,
    BarangHabisPakaiMasuk,
    BarangHabisPakaiKeluar,
)

# =========================================================
# DASHBOARD
# =========================================================
def dashboard(request):
    # =====================================================
    # TOTAL ASET & RINGKASAN
    # =====================================================
    total_aset = sum([
        Tanah.objects.count(),
        PeralatanMesin.objects.count(),
        Gedung.objects.count(),
        Jalan.objects.count(),
        Buku.objects.count(),
    ])

    total_bhp = BarangHabisPakai.objects.count()
    peminjaman_aktif = Peminjaman.objects.filter(status='dipinjam').count()
    stok_habis = BarangHabisPakai.objects.filter(stok=0).count()

    # =====================================================
    # GRAFIK PERTUMBUHAN ASET (PER TAHUN)
    # =====================================================
    aset_per_tahun = (
        PeralatanMesin.objects
        .values("tahun_perolehan")
        .annotate(total=Count("id"))
        .order_by("tahun_perolehan")
    )

    labels_aset = [str(item["tahun_perolehan"]) for item in aset_per_tahun]
    data_aset = [item["total"] for item in aset_per_tahun]

    # ======================================================
    # GRAFIK PEMINJAMAN (PER BULAN)
    # ======================================================
    pinjam_per_bulan = (
        Peminjaman.objects
        .annotate(bulan=TruncMonth("tanggal_pinjam"))
        .values("bulan")
        .annotate(total=Count("id"))
        .order_by("bulan")
    )

    labels_pinjam = [
        item["bulan"].strftime("%b %Y")
        for item in pinjam_per_bulan
        if item["bulan"]
    ]

    data_pinjam = [
        item["total"]
        for item in pinjam_per_bulan
        if item["bulan"]
    ]

    # ======================================================
    # CONTEXT
    # ======================================================
    context = {
        "total_aset": total_aset,
        "total_bhp": total_bhp,
        "peminjaman_aktif": peminjaman_aktif,
        "stok_habis": stok_habis,
        "labels_aset": json.dumps(labels_aset),
        "data_aset": json.dumps(data_aset),
        "labels_pinjam": json.dumps(labels_pinjam),
        "data_pinjam": json.dumps(data_pinjam),
    }

    return render(request, "sarpras/dashboard.html", context)

# ===============================================================
# KIB A - TANAH
# ===============================================================
def tanah_list(request):
    data = Tanah.objects.all().order_by('kode_barang')
    return render(request, 'sarpras/tanah.html', {'data': data})

def tanah_tambah(request):
    if request.method == "POST":
        Tanah.objects.create(
            kode_barang=request.POST.get('kode_barang'),
            nama=request.POST.get('nama'),
            luas=float(request.POST.get('luas')),
            lokasi=request.POST.get('lokasi'),
            status=request.POST.get('status'),
            tahun_perolehan=int(request.POST.get('tahun_perolehan')),
        )
        return redirect('tanah_list')

    return render(request, 'sarpras/tanah_form.html')

def tanah_edit(request, pk):
    data = get_object_or_404(Tanah, pk=pk)

    if request.method == "POST":
        data.kode_barang = request.POST.get('kode_barang')
        data.nama = request.POST.get('nama')
        data.luas = float(request.POST.get('luas'))
        data.lokasi = request.POST.get('lokasi')
        data.status = request.POST.get('status')
        data.tahun_perolehan = int(request.POST.get('tahun_perolehan'))
        data.save()

        return redirect('tanah_list')

    return render(request, 'sarpras/tanah_form.html', {
        'data': data,
        'edit': True
    })

def tanah_hapus(request, pk):
    tanah = get_object_or_404(Tanah, pk=pk)
    tanah.delete()
    return redirect('tanah_list')

#================================================================
#CETAK TANAH PDF
#================================================================


def tanah_cetak_pdf(request):
    data = Tanah.objects.all().order_by('nama')

    total_bidang = data.count()
    total_luas = data.aggregate(total=Sum('luas'))['total'] or 0

    # ================= LOGO =================
    logo_path = os.path.join(
        settings.BASE_DIR,
        'sarpras',
        'static',
        'sarpras',
        'logo_sekolah.png'
    )

    with open(logo_path, "rb") as image_file:
        logo_base64 = base64.b64encode(image_file.read()).decode()

    # ================= QR CODE =================
    qr_text = (
        "DOKUMEN RESMI\n"
        "Laporan Tanah KIB A\n"
        f"Tanggal Cetak: {timezone.now().strftime('%d-%m-%Y %H:%M')}\n"
        "Sumber: Sistem SARPRAS"
    )

    qr = qrcode.make(qr_text)
    buffer = BytesIO()
    qr.save(buffer, format="PNG")
    qr_base64 = base64.b64encode(buffer.getvalue()).decode()

    context = {
        'data': data,
        'tahun': timezone.now().year,
        'tanggal': timezone.now(),
        'lokasi': 'Lubuklinggau',
        'qr_code': qr_base64,
        'logo_base64': logo_base64,
        'total_bidang': total_bidang,
        'total_luas': total_luas,
    }

    template = get_template("sarpras/tanah_cetak_pdf.html")
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="laporan_tanah.pdf"'

    pisa.CreatePDF(html, dest=response)

    return response

# ===================================================================
# KIB B - PERALATAN & MESIN
# ===================================================================

def peralatan_list(request):
    query = request.GET.get('q', '')
    gedung_id = request.GET.get('gedung')

    qs = PeralatanMesin.objects.select_related('ruangan__gedung').all().order_by('-id')

    # 🔎 Filter pencarian
    if query:
        qs = qs.filter(
            Q(kode_barang__icontains=query) |
            Q(nama__icontains=query) |
            Q(kondisi__icontains=query)
        )

    # 🏢 Filter berdasarkan gedung
    if gedung_id:
        qs = qs.filter(ruangan__gedung_id=gedung_id)

    paginator = Paginator(qs, 5)
    page_number = request.GET.get('page')
    peralatan = paginator.get_page(page_number)

    return render(request, 'sarpras/peralatan.html', {
        'data': peralatan,
        'peralatan': peralatan,
        'query': query,
        'gedungs': Gedung.objects.all(),
        'selected_gedung': gedung_id,
    })


def peralatan_tambah(request):
    gedungs = Gedung.objects.all()
    ruangans = Ruangan.objects.all()

    if request.method == 'POST':
        ruangan_id = request.POST.get('ruangan')

        PeralatanMesin.objects.create(
            kode_barang=request.POST.get('kode_barang'),
            nama=request.POST.get('nama'),
            jumlah=request.POST.get('jumlah'),
            kondisi=request.POST.get('kondisi'),
            tahun_perolehan=request.POST.get('tahun_perolehan'),
            gambar=request.FILES.get('gambar'),
            ruangan_id=ruangan_id if ruangan_id else None
        )

        return redirect('peralatan_list')

    return render(request, 'sarpras/peralatan_form.html', {
        'gedungs': gedungs,
        'ruangans': ruangans
    })


def peralatan_edit(request, id):
    data = get_object_or_404(PeralatanMesin, id=id)

    if request.method == 'POST':
        data.kode_barang = request.POST.get('kode_barang')
        data.nama = request.POST.get('nama')
        data.jumlah = request.POST.get('jumlah')
        data.kondisi = request.POST.get('kondisi')
        data.tahun_perolehan = request.POST.get('tahun_perolehan')

        if request.FILES.get('gambar'):
            data.gambar = request.FILES.get('gambar')

        data.save()
        return redirect('peralatan_list')

    return render(
        request,
        'sarpras/peralatan_form.html',
        {'data': data,
         'edit': True
         
         })

def peralatan_hapus(request, id):
    data = get_object_or_404(PeralatanMesin, id=id)
    data.delete()
    return redirect('peralatan_list')


#Import peralatan========================

def peralatan_import(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        wb = openpyxl.load_workbook(file)
        sheet = wb.active

        gagal = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                if not row[0]:
                    continue

                PeralatanMesin.objects.create(
                    kode_barang=str(row[0]).strip(),
                    nama=str(row[1]).strip(),
                    jumlah=int(float(row[2])) if row[2] else 0,                                    
                    kondisi=str(row[4]).strip(),
                    tahun_perolehan=int(float(row[4])) if row[4] else 0,
                )

            except Exception:
                gagal += 1
                continue

        messages.success(
            request,
            f'Import selesai. Data gagal: {gagal}'
        )
        return redirect('peralatan_list')

    return render(request, 'sarpras/peralatan_import.html')


#export Peralatan==================================


def peralatan_export_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "KIB B - Peralatan & Mesin"

    # =====================
    # HEADER
    # =====================
    headers = [
        "No",
        "Kode Barang",
        "Nama Peralatan",
        "Jumlah",
        "Kondisi",
        "Tahun Perolehan"
        
    ]

    ws.append(headers)

    # Style header
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # =====================
    # DATA
    # =====================
    data = PeralatanMesin.objects.all().order_by('nama')

    for i, peralatan in enumerate(data, start=1):
        ws.append([
            i,
            peralatan.kode_barang,
            peralatan.nama,
            peralatan.jumlah,
            peralatan.kondisi,
            peralatan.tahun_perolehan,
           
        ])

    # =====================
    # RESPONSE
    # =====================
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="kib_B_Peralatan.xlsx"'

    wb.save(response)
    return response

# Rekap KIB B ==========================================

def peralatan_rekap(request):
    data = (
        PeralatanMesin.objects
        .values('nama')
        .annotate(total=Sum('jumlah'))
        .order_by('nama')
    )

    total_semua = (
        PeralatanMesin.objects
        .aggregate(total=Sum('jumlah'))
        ['total'] or 0
    )

    return render(
        request,
        'sarpras/peralatan_rekap.html',
        {
            'data': data,
            'total_semua': total_semua
        }
    )

#cetak peralatan pdf =================================

def peralatan_cetak_pdf(request):
    data = PeralatanMesin.objects.all().order_by('nama')

    total_jenis = data.count()
    total_unit = data.aggregate(total=Sum('jumlah'))['total'] or 0

    # ================= LOGO BASE64 =================
    logo_path = os.path.join(
        settings.BASE_DIR,
        'sarpras',
        'static',
        'sarpras',
        'logo_sekolah.png'
    )

    with open(logo_path, "rb") as image_file:
        logo_base64 = base64.b64encode(image_file.read()).decode()

    # ================= QR CODE =================
    qr_text = (
        "DOKUMEN RESMI\n"
        "Laporan Peralatan KIB B\n"
        f"Tanggal Cetak: {timezone.now().strftime('%d-%m-%Y %H:%M')}\n"
        "Sumber: Sistem SARPRAS"
    )

    qr = qrcode.make(qr_text)
    buffer = BytesIO()
    qr.save(buffer, format="PNG")
    qr_base64 = base64.b64encode(buffer.getvalue()).decode()

    context = {
        "data": data,
        "logo_base64": logo_base64,
        "qr_code": qr_base64,
        "total_jenis": total_jenis,
        "total_unit": total_unit,
        "tahun": timezone.now().year,
        "tanggal": timezone.now(),
        "lokasi": "Lubuklinggau",
    }

    template = get_template("sarpras/peralatan_cetak_pdf.html")
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="laporan_peralatan.pdf"'

    pisa.CreatePDF(html, dest=response)

    return response

# =========================
# KIB C - GEDUNG & BANGUNAN
# =========================
def gedung_list(request):
    data = Gedung.objects.all().order_by('kode_barang')
    return render(request, 'sarpras/gedung.html', {'data': data})

def gedung_tambah(request):
    if request.method == 'POST':
        Gedung.objects.create(
            kode_barang=request.POST.get('kode_barang'),
            nama=request.POST.get('nama'),
            lokasi=request.POST.get('lokasi'),
            luas=request.POST.get('luas'),
            kondisi=request.POST.get('kondisi'),
            tahun_perolehan=request.POST.get('tahun_perolehan')
        )
        return redirect('/gedung/')

    return render(request, 'sarpras/gedung_form.html')



def gedung_edit(request, pk):
    gedung = get_object_or_404(Gedung, pk=pk)

    if request.method == 'POST':
        gedung.kode_barang = request.POST.get('kode_barang')
        gedung.nama = request.POST.get('nama')
        gedung.lokasi = request.POST.get('lokasi')
        gedung.luas = request.POST.get('luas')
        gedung.kondisi = request.POST.get('kondisi')
        gedung.tahun_perolehan = request.POST.get('tahun_perolehan')
        gedung.save()
        return redirect('gedung_list')

    return render(request, 'sarpras/gedung_form.html', {'gedung': gedung})

def gedung_hapus(request, pk):
    gedung = get_object_or_404(Gedung, pk=pk)
    gedung.delete()
    return redirect('gedung_list')


#cetak gedung pdf

def gedung_cetak_pdf(request):
    data = Gedung.objects.all().order_by('nama')

    total_gedung = data.count()
    total_luas = data.aggregate(total=Sum('luas'))['total'] or 0

    # ================= LOGO BASE64 =================
    logo_path = os.path.join(
        settings.BASE_DIR,
        'sarpras',
        'static',
        'sarpras',
        'logo_sekolah.png'
    )

    with open(logo_path, "rb") as image_file:
        logo_base64 = base64.b64encode(image_file.read()).decode()

    # ================= QR CODE =================
    qr_text = (
        "DOKUMEN RESMI\n"
        "Laporan Gedung KIB C\n"
        f"Tanggal Cetak: {timezone.now().strftime('%d-%m-%Y %H:%M')}\n"
        "Sumber: Sistem SARPRAS"
    )

    qr = qrcode.make(qr_text)
    buffer = BytesIO()
    qr.save(buffer, format="PNG")
    qr_base64 = base64.b64encode(buffer.getvalue()).decode()

    context = {
        'data': data,
        'tahun': timezone.now().year,
        'tanggal': timezone.now(),
        'lokasi': 'Lubuklinggau',
        'qr_code': qr_base64,
        'logo_base64': logo_base64,   # 🔥 WAJIB
        'total_gedung': total_gedung,
        'total_luas': total_luas,
    }

    template = get_template("sarpras/gedung_cetak_pdf.html")
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="laporan_gedung.pdf"'

    pisa.CreatePDF(html, dest=response)

    return response

# =====================
# KIB D - JALAN / IRIGASI / JARINGAN
# =====================
def jalan_list(request):
    data = Jalan.objects.all().order_by('kode_barang')
    return render(request, 'sarpras/jalan.html', {'data': data})

def jalan_tambah(request):
    if request.method == 'POST':
        Jalan.objects.create(
            kode_barang=request.POST.get('kode_barang'),
            nama=request.POST.get('nama'),
            panjang=request.POST.get('panjang'),
            lokasi=request.POST.get('lokasi'),
            kondisi=request.POST.get('kondisi'),
            tahun_perolehan=request.POST.get('tahun_perolehan')
        )
        return redirect('/jalan/')

    return render(request, 'sarpras/jalan_form.html')

def jalan_edit(request, pk):
    jalan = get_object_or_404(Jalan, pk=pk)

    if request.method == 'POST':
        jalan.kode_barang = request.POST.get('kode_barang')
        jalan.nama = request.POST.get('nama')
        jalan.panjang = request.POST.get('panjang')
        jalan.lokasi = request.POST.get('lokasi')
        jalan.kondisi = request.POST.get('kondisi')
        jalan.tahun_perolehan = request.POST.get('tahun_perolehan')
        jalan.save()
        return redirect('jalan_list')

    return render(request, 'sarpras/jalan_form.html', {'jalan': jalan})

def jalan_hapus(request, pk):
    jalan = get_object_or_404(Jalan, pk=pk)
    jalan.delete()
    return redirect('jalan_list')

#cetak jalan pdf ====================================


def jalan_cetak_pdf(request):
    data = Jalan.objects.all().order_by('nama')

    total_jalan = data.count()
    total_panjang = data.aggregate(total=Sum('panjang'))['total'] or 0

    # ================= LOGO BASE64 =================
    logo_path = os.path.join(
        settings.BASE_DIR,
        'sarpras',
        'static',
        'sarpras',
        'logo_sekolah.png'
    )

    with open(logo_path, "rb") as image_file:
        logo_base64 = base64.b64encode(image_file.read()).decode()

    # ================= QR CODE =================
    qr_text = (
        "DOKUMEN RESMI\n"
        "Laporan Jalan KIB D\n"
        f"Tanggal Cetak: {timezone.now().strftime('%d-%m-%Y %H:%M')}\n"
        "Sumber: Sistem SARPRAS"
    )

    qr = qrcode.make(qr_text)
    buffer = BytesIO()
    qr.save(buffer, format="PNG")
    qr_base64 = base64.b64encode(buffer.getvalue()).decode()

    context = {
        'data': data,
        'tahun': timezone.now().year,
        'tanggal': timezone.now(),
        'lokasi': 'Lubuklinggau',
        'qr_code': qr_base64,
        'logo_base64': logo_base64,
        'total_jalan': total_jalan,
        'total_panjang': total_panjang,
    }

    template = get_template("sarpras/jalan_cetak_pdf.html")
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="laporan_jalan.pdf"'

    pisa.CreatePDF(html, dest=response)

    return response

# ==============================================================================
# KIB E - BUKU
# ==============================================================================

def buku_list(request):
    query = request.GET.get('q', '')

    buku_qs = Buku.objects.all().order_by('-id')

    if query:
        buku_qs = buku_qs.filter(
            Q(judul__icontains=query) |
            Q(kode_barang__icontains=query) |
            Q(pengarang__icontains=query)
        )

    paginator = Paginator(buku_qs, 10)  # ⬅️ 10 buku per halaman
    page_number = request.GET.get('page')
    buku = paginator.get_page(page_number)

    context = {
        'buku': buku,
        'query': query,
    }
    return render(request, 'sarpras/buku.html', context)

def buku_tambah(request):
    if request.method == 'POST':
        Buku.objects.create(
            kode_barang=request.POST.get('kode_barang'),
            judul=request.POST.get('judul'),
            pengarang=request.POST.get('pengarang'),
            jumlah=request.POST.get('jumlah'),
            kondisi=request.POST.get('kondisi'),
            tahun_terbit=request.POST.get('tahun_terbit'),
            gambar=request.FILES.get('gambar')
        )
        return redirect('/buku/')

    return render(request, 'sarpras/buku_form.html')

def buku_edit(request, pk):
    buku = get_object_or_404(Buku, pk=pk)

    if request.method == 'POST':
        buku.kode_barang = request.POST.get('kode_barang')
        buku.judul = request.POST.get('judul')
        buku.pengarang = request.POST.get('pengarang')
        buku.jumlah = request.POST.get('jumlah')
        buku.kondisi = request.POST.get('kondisi')
        buku.tahun_terbit = request.POST.get('tahun_terbit')

        if request.FILES.get('gambar'):
            buku.gambar = request.FILES.get('gambar')

        buku.save()
        return redirect('buku_list')

    return render(request, 'sarpras/buku_form.html', {
        'buku': buku,
        'edit': True
        })


def buku_hapus(request, pk):
    buku = get_object_or_404(Buku, pk=pk)
    buku.delete()
    return redirect('buku_list')


# KIB E - REKAP BUKU ================================================================

def buku_rekap(request):
    # ======================
    # AMBIL KEYWORD
    # ======================
    keyword = request.GET.get('q', '').strip()

    buku_qs = Buku.objects.all()

    if keyword:
        buku_qs = buku_qs.filter(judul__icontains=keyword)

    # ======================
    # RINGKASAN
    # ======================
    total_judul = buku_qs.values('judul').distinct().count()
    total_eksemplar = buku_qs.aggregate(
        total=Sum('jumlah')
    )['total'] or 0

    # ======================
    # REKAP KONDISI
    # ======================
    rekap_kondisi = (
        buku_qs
        .values('kondisi')
        .annotate(total=Sum('jumlah'))
        .order_by('kondisi')
    )

    # ======================
    # REKAP PER JUDUL
    # ======================
    rekap_judul = (
        buku_qs
        .values('judul')
        .annotate(total=Sum('jumlah'))
        .order_by('-total')
    )

    # ======================
    # PAGINATION (SEARCH-AWARE)
    # ======================
    paginator = Paginator(rekap_judul, 10)  # 10 judul / halaman
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'total_judul': total_judul,
        'total_eksemplar': total_eksemplar,
        'rekap_kondisi': rekap_kondisi,
        'page_obj': page_obj,
        'keyword': keyword,
    }

    return render(request, 'sarpras/buku_rekap.html', context)

# ======================
# Inport Buku Exel
# ======================
def buku_import(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        wb = openpyxl.load_workbook(file)
        sheet = wb.active

        gagal = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                if not row[0]:
                    continue

                Buku.objects.create(
                    kode_barang=str(row[0]).strip(),
                    judul=str(row[1]).strip(),
                    pengarang=str(row[2]).strip(),
                    jumlah=int(float(row[3])) if row[3] else 0,
                    kondisi=str(row[4]).strip(),
                    tahun_terbit=int(float(row[5])) if row[5] else 0,
                )

            except Exception as e:
                gagal += 1
                continue

        messages.success(
            request,
            f'Import selesai. Data gagal: {gagal}'
        )
        return redirect('buku_list')

    return render(request, 'sarpras/buku_import.html')

#===================================
#export buku exel
#====================================

def buku_export_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "KIB E - Buku"

    # =====================
    # HEADER
    # =====================
    headers = [
        "No",
        "Kode Buku",
        "Judul Buku",
        "Pengarang",
        "Jumlah",
        "Kondisi",
        "Tahun Terbit"
    ]

    ws.append(headers)

    # Style header
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # =====================
    # DATA
    # =====================
    data = Buku.objects.all().order_by('judul')

    for i, buku in enumerate(data, start=1):
        ws.append([
            i,
            buku.kode_barang,
            buku.judul,
            buku.pengarang,
            buku.jumlah,
            buku.kondisi,
            buku.tahun_terbit,
        ])

    # =====================
    # RESPONSE
    # =====================
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="kib_e_buku.xlsx"'

    wb.save(response)
    return response


# cetak pdf ===============================

def buku_cetak_pdf(request):
    data = Buku.objects.all().order_by('judul')

    # =========================
    # AMBIL LOGO DARI STATIC → BASE64 (PALING STABIL)
    # =========================
    from django.contrib.staticfiles import finders
    import base64
    from io import BytesIO
    import qrcode

    logo_path = finders.find('sarpras/logo_sekolah.png')

    if not logo_path:
        raise Exception("Logo tidak ditemukan di static!")

    with open(logo_path, "rb") as img_file:
        logo_base64 = base64.b64encode(img_file.read()).decode()

    # =========================
    # QR CODE
    # =========================
    qr_text = (
        "DOKUMEN RESMI\n"
        "Laporan Buku KIB E\n"
        f"Tanggal Cetak: {timezone.now().strftime('%d-%m-%Y %H:%M')}\n"
        "Sumber: Sistem SARPRAS"
    )

    qr = qrcode.make(qr_text)
    buffer = BytesIO()
    qr.save(buffer, format="PNG")
    qr_base64 = base64.b64encode(buffer.getvalue()).decode()

    # =========================
    # RENDER TEMPLATE
    # =========================
    template = get_template('sarpras/buku_cetak_pdf.html')
    html = template.render({
        'data': data,
        'tahun': timezone.now().year,
        'tanggal': timezone.now(),
        'lokasi': 'Lubuklinggau',
        'qr_code': qr_base64,
        'logo_base64': logo_base64,
    })

    # =========================
    # GENERATE PDF
    # =========================
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="laporan_kib_e_buku.pdf"'

    pisa.CreatePDF(html, dest=response)

    return response


    # =========================
    # RESPONSE PDF
    # =========================
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="laporan_kib_e_buku.pdf"'

    # 🔥 LINK CALLBACK (INI KUNCI UTAMA)
    def link_callback(uri, rel):
        if uri.startswith('file:///'):
            return uri.replace('file:///', '')
        return uri

    pisa.CreatePDF(
        html,
        dest=response,
        link_callback=link_callback
    )

    return response

#=============================
# Cetak Semua KIB PDF
#=============================

def semua_kib_cetak_pdf(request):
    qr = qrcode.make("Laporan Lengkap Aset Sekolah | Sistem SARPRAS")
    buffer = BytesIO()
    qr.save(buffer, format="PNG")
    qr_base64 = base64.b64encode(buffer.getvalue()).decode()

    template = get_template('sarpras/semua_kib_cetak_pdf.html')
    html = template.render({
        'tanah': Tanah.objects.all(),
        'peralatan': PeralatanMesin.objects.all(),
        'gedung': Gedung.objects.all(),
        'jalan': Jalan.objects.all(),
        'buku': Buku.objects.all(),
        'tanggal': timezone.now(),
        'qr_code': qr_base64,
    })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="laporan_semua_kib.pdf"'
    pisa.CreatePDF(html, dest=response)
    return response


# =====================================================
# LIST PEMINJAMAN
# =====================================================
def peminjaman_list(request):
    data = Peminjaman.objects.all().order_by('-tanggal_pinjam')
    return render(request, 'peminjaman/list.html', {'data': data})


# =====================================================
# TAMBAH PEMINJAMAN (KURANGI STOK SEKALI SAJA)
# =====================================================
def peminjaman_create(request):
    if request.method == 'POST':
        with transaction.atomic():
            barang = get_object_or_404(
                PeralatanMesin.objects.select_for_update(),
                id=request.POST['barang']
            )

            jumlah = int(request.POST['jumlah_pinjam'])

            # VALIDASI
            if jumlah <= 0:
                return render(request, 'peminjaman/form.html', {
                    'barang': PeralatanMesin.objects.all(),
                    'error': 'Jumlah pinjam harus lebih dari 0'
                })

            if barang.jumlah < jumlah:
                return render(request, 'peminjaman/form.html', {
                    'barang': PeralatanMesin.objects.all(),
                    'error': f'Stok tidak cukup. Sisa {barang.jumlah}'
                })

            # 🔥 SATU-SATUNYA TEMPAT KURANGI STOK
            barang.jumlah = barang.jumlah - jumlah
            barang.save(update_fields=['jumlah'])

            # CATAT PEMINJAMAN (TIDAK MENGUBAH STOK)
            Peminjaman.objects.create(
                barang=barang,
                peminjam=request.POST['peminjam'],
                jumlah_pinjam=jumlah,
                tanggal_kembali=request.POST.get('tanggal_kembali'),
                status='dipinjam'
            )

        return redirect('peminjaman_list')

    return render(request, 'peminjaman/form.html', {
        'barang': PeralatanMesin.objects.all()
    })


# =====================================================
# LIST BARANG YANG MASIH DIPINJAM
# =====================================================
def pengembalian_list(request):
    data = Peminjaman.objects.filter(status='dipinjam').order_by('-tanggal_pinjam')
    return render(request, 'peminjaman/pengembalian_list.html', {
        'data': data
    })


# =====================================================
# KONFIRMASI PENGEMBALIAN (TAMBAH STOK SEKALI SAJA)
# =====================================================
def peminjaman_kembali(request, id):
    with transaction.atomic():
        pinjam = get_object_or_404(
            Peminjaman.objects.select_for_update(),
            id=id
        )

        # CEGAH KLIK GANDA
        if pinjam.status == 'kembali':
            return redirect('pengembalian_list')

        barang = pinjam.barang

        # 🔥 SATU-SATUNYA TEMPAT TAMBAH STOK
        barang.jumlah = barang.jumlah + pinjam.jumlah_pinjam
        barang.save(update_fields=['jumlah'])

        pinjam.status = 'kembali'
        pinjam.tanggal_kembali = timezone.now().date()
        pinjam.save(update_fields=['status', 'tanggal_kembali'])

    return redirect('pengembalian_list')

# =====================================================
# CETAK SURAT PEMINJAMAN PERALATAN (PDF RESMI)
# =====================================================
def cetak_surat_peminjaman(request, id):
    # 1. Ambil data peminjaman
    p = get_object_or_404(Peminjaman, id=id)

    # 2. Data statis sekolah (boleh nanti pindah ke setting / model)
    context = {
        'p': p,
        'sekolah': 'Kabupaten XXXXX',
        'dinas': 'Dinas Pendidikan Kabupaten XXXXX',
        'nama_sekolah': 'SMKN XXXXX',
        'alamat_sekolah': 'Jl. XXXXX No. XX',
        'kepala_sekolah': 'Nama Kepala Sekolah',
        'nomor_surat': f'421.5/{p.id}/SARPRAS/{p.tanggal_pinjam.year}',
        'tanggal_hari_ini': timezone.now().date(),
    }

    # 3. Render HTML → string
    html = render_to_string(
        'peminjaman/surat_peminjaman.html',
        context
    )

    # 4. Generate PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = (
        f'inline; filename="surat_peminjaman_{p.id}.pdf"'
    )

    pisa.CreatePDF(html, dest=response)

    return response

# ===========================================================
# BARANG HABIS PAKAI (BHP)
# ===========================================================

def bhp_list(request):
    q = request.GET.get('q', '')

    data = BarangHabisPakai.objects.all().order_by('nama_barang')

    if q:
        data = data.filter(
            Q(nama_barang__icontains=q) |
            Q(kode_barang__icontains=q)
        )

    return render(request, 'bhp/list.html', {
        'data': data,
        'q': q,
    })



def bhp_tambah(request):
    if request.method == 'POST':
        BarangHabisPakai.objects.create(
            kode_barang=request.POST['kode_barang'],
            nama_barang=request.POST['nama_barang'],
            satuan=request.POST['satuan'],
            stok=int(request.POST.get('stok', 0))
        )
        messages.success(
            request,
            'Barang habis pakai berhasil ditambahkan'
        )
        return redirect('bhp_list')

    return render(request, 'bhp/tambah.html')


def bhp_masuk(request):
    if request.method == 'POST':
        with transaction.atomic():
            barang = get_object_or_404(
                BarangHabisPakai.objects.select_for_update(),
                id=request.POST['barang']
            )

            jumlah = int(request.POST['jumlah'])

            barang.stok += jumlah
            barang.save(update_fields=['stok'])

            BarangHabisPakaiMasuk.objects.create(
                barang=barang,
                jumlah=jumlah,
                sumber=request.POST['sumber'],
                keterangan=request.POST.get('keterangan', '')
            )

        messages.success(
            request,
            'Barang masuk berhasil dicatat'
        )
        return redirect('bhp_list')

    return render(request, 'bhp/masuk.html', {
        'barang': BarangHabisPakai.objects.all()
    })


def bhp_keluar(request):
    if request.method == 'POST':
        with transaction.atomic():
            barang = get_object_or_404(
                BarangHabisPakai.objects.select_for_update(),
                id=request.POST['barang']
            )

            jumlah = int(request.POST['jumlah'])

            if barang.stok < jumlah:
                messages.error(
                    request,
                    f"Stok tidak cukup! Sisa stok: {barang.stok}"
                )
                return redirect('bhp_keluar')

            barang.stok -= jumlah
            barang.save(update_fields=['stok'])

            BarangHabisPakaiKeluar.objects.create(
                barang=barang,
                jumlah=jumlah,
                pengguna=request.POST['pengguna'],
                keperluan=request.POST['keperluan'],
                keterangan=request.POST.get('keterangan', '')
            )

        messages.success(
            request,
            'Barang keluar berhasil dicatat'
        )
        return redirect('bhp_list')

    return render(request, 'bhp/keluar.html', {
        'barang': BarangHabisPakai.objects.all()
    })


#======================================
#Import Habis Pakai
#========================================
def bhp_import(request):
    if request.method != 'POST':
        return redirect('bhp_list')

    file = request.FILES.get('file')

    if not file:
        messages.error(request, 'File belum dipilih')
        return redirect('bhp_list')

    try:
        # ===============================
        # IMPORT CSV
        # ===============================
        if file.name.endswith('.csv'):
            csv_file = TextIOWrapper(file.file, encoding='utf-8')
            reader = csv.DictReader(csv_file)

            for row in reader:
                if not row.get('kode') or not row.get('nama_barang'):
                    continue

                try:
                    stok = int(row.get('stok', 0))
                except ValueError:
                    stok = 0

                BarangHabisPakai.objects.update_or_create(
                    kode_barang=row['kode'],
                    defaults={
                        'nama_barang': row['nama_barang'],
                        'stok': stok,
                        'satuan': row.get('satuan', '')
                    }
                )

        # ===============================
        # IMPORT EXCEL
        # ===============================
        elif file.name.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                kode, nama, stok, satuan = row

                # Lewati baris kosong
                if not kode or not nama:
                    continue

                # Amankan stok
                try:
                    stok = int(stok) if stok is not None else 0
                except ValueError:
                    stok = 0

                BarangHabisPakai.objects.update_or_create(
                    kode_barang=str(kode),
                    defaults={
                        'nama_barang': nama,
                        'stok': stok,
                        'satuan': satuan if satuan else ''
                    }
                )

        else:
            messages.error(request, 'Format file harus CSV atau XLSX')
            return redirect('bhp_list')

        messages.success(request, 'Import barang habis pakai berhasil')

    except Exception as e:
        messages.error(request, f'Gagal import: {e}')

    return redirect('bhp_list')


#========================================
# transaksi bhp
#========================================

def bhp_transaksi(request):
    masuk = BarangHabisPakaiMasuk.objects.select_related('barang').order_by('-tanggal')
    keluar = BarangHabisPakaiKeluar.objects.select_related('barang').order_by('-tanggal')

    return render(request, 'bhp/transaksi.html', {
        'masuk': masuk,
        'keluar': keluar
    })




def bhp_transaksi_pdf(request):
    bulan = request.GET.get('bulan')
    tahun = request.GET.get('tahun')

    masuk = BarangHabisPakaiMasuk.objects.select_related('barang')
    keluar = BarangHabisPakaiKeluar.objects.select_related('barang')

    judul_periode = "Semua Periode"

    if bulan and tahun:
        masuk = masuk.filter(tanggal__month=bulan, tanggal__year=tahun)
        keluar = keluar.filter(tanggal__month=bulan, tanggal__year=tahun)

        nama_bulan = calendar.month_name[int(bulan)]
        judul_periode = f"{nama_bulan} {tahun}"

    masuk = masuk.order_by('tanggal')
    keluar = keluar.order_by('tanggal')

    template = get_template('bhp/transaksi_pdf.html')
    html = template.render({
        'masuk': masuk,
        'keluar': keluar,
        'tanggal_cetak': date.today(),
        'periode': judul_periode
    })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="laporan_bhp.pdf"'

    pisa.CreatePDF(html, dest=response)
    return response



#riwayat habis pakai ===========


def bhp_barang_riwayat(request, barang_id):
    barang = get_object_or_404(BarangHabisPakai, id=barang_id)

    masuk = BarangHabisPakaiMasuk.objects.filter(
        barang=barang
    ).order_by('-tanggal')

    keluar = BarangHabisPakaiKeluar.objects.filter(
        barang=barang
    ).order_by('-tanggal')

    total_masuk = masuk.aggregate(
        total=Sum('jumlah')
    )['total'] or 0

    total_keluar = keluar.aggregate(
        total=Sum('jumlah')
    )['total'] or 0

    context = {
        'barang': barang,
        'masuk': masuk,
        'keluar': keluar,
        'total_masuk': total_masuk,
        'total_keluar': total_keluar,
        'stok_sisa': barang.stok,
    }

    return render(
        request,
        'bhp/riwayat_barang.html',
        context
    )


#ruangan baru=================


def aset_per_ruangan(request):
    gedungs = Gedung.objects.all()
    return render(request, 'sarpras/aset_per_ruangan.html', {
        'gedungs': gedungs
    })


# =====================
# RUANGAN
# =====================


def ruangan_list(request):
    data = Ruangan.objects.select_related('gedung').all().order_by('gedung__nama', 'nama')
    return render(request, 'sarpras/ruangan.html', {'data': data})


def ruangan_tambah(request):
    if request.method == "POST":
        Ruangan.objects.create(
            gedung_id=request.POST.get('gedung'),
            nama=request.POST.get('nama'),
            kode=request.POST.get('kode'),
            penanggung_jawab=request.POST.get('penanggung_jawab'),
        )
        return redirect('ruangan_list')

    gedungs = Gedung.objects.all()
    return render(request, 'sarpras/ruangan_form.html', {
        'gedungs': gedungs
    })


def ruangan_edit(request, pk):
    ruangan = get_object_or_404(Ruangan, pk=pk)

    if request.method == "POST":
        ruangan.gedung_id = request.POST.get('gedung')
        ruangan.nama = request.POST.get('nama')
        ruangan.kode = request.POST.get('kode')
        ruangan.penanggung_jawab = request.POST.get('penanggung_jawab')
        ruangan.save()
        return redirect('ruangan_list')

    gedungs = Gedung.objects.all()
    return render(request, 'sarpras/ruangan_form.html', {
        'data': ruangan,
        'gedungs': gedungs,
        'edit': True
    })


def ruangan_hapus(request, pk):
    ruangan = get_object_or_404(Ruangan, pk=pk)
    ruangan.delete()
    return redirect('ruangan_list')



#============Cetak KIR ==================

def cetak_kir(request, id):
    ruangan = get_object_or_404(Ruangan, id=id)
    peralatan = ruangan.peralatan.all()

    total_unit = sum([p.jumlah for p in peralatan])

    return render(request, 'sarpras/kartu_inventaris_ruangan.html', {
        'ruangan': ruangan,
        'peralatan': peralatan,
        'total_unit': total_unit
    })
